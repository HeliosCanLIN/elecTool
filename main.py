from selenium import webdriver #实例化浏览器
from selenium.webdriver import ActionChains #浏览器自动操作
import pyperclip #获取剪切板
import win32api #非input标签上传文件需要的系统库
import win32con #非input标签上传文件需要的系统库
from openpyxl import load_workbook #读取Excel文件
from openpyxl.utils import get_column_letter, column_index_from_string #操作Excel文件
import datetime
import time
import json #读取json文件
import sys
import yaml #读取yaml配置文件
import requests #发送报文请求
import re #正则表达式
from bs4 import BeautifulSoup #处理html
import queue #队列

version = "v0.8.3-beta"
versionNum = "0824.2"

def get_yaml_data(yaml_file): #读取配置文件
    # 打开yaml文件
    print("***获取配置文件数据***")
    file = open(yaml_file, 'r', encoding="utf-8")
    file_data = file.read()
    file.close()
    data = yaml.safe_load(file_data)
    return data

def new_round(_float, _len): #重写的四舍五入功能
    """
    Parameters
    ----------
    _float: float
    _len: int, 指定四舍五入需要保留的小数点后几位数为_len

    Returns
    -------
    type ==> float, 返回四舍五入后的值
    """
    if isinstance(_float, float):
        if str(_float)[::-1].find('.') <= _len:
            return (_float)
        if str(_float)[-1] == '5':
            return (round(float(str(_float)[:-1] + '6'), _len))
        else:
            return (round(_float, _len))
    else:
        return (round(_float, _len))

def UpLoad_File(webEle, filePath): #非ipunt标签的上传附件功能
    """
    非input标签的文件上传功能 (定位好的元素,文件路径)
    使用 python 的 win32api，win32con 模拟按键输入，实现文件上传操作。
    :param webEle: 页面中的上传文件按钮,是已经获取到的对象
    :param filePath: 要上传的文件地址，绝对路径。如：D:\\timg (1).jpg
    :param check_Input:检查input标签中是否有值 #仅用来检查，在return 处调用一次，多余可删除
    """
    pyperclip.copy(filePath)  # 复制文件路径到剪切板
    ActionChains(browser).move_to_element(webEle).perform()
    webEle.click()  # 点击上传图片按钮
    time.sleep(0.5)  # 等待程序加载 时间 看你电脑的速度 单位(秒)
    # 发送 ctrl（17） + V（86）按钮
    win32api.keybd_event(17, 0, 0, 0)
    win32api.keybd_event(86, 0, 0, 0)
    win32api.keybd_event(86, 0, win32con.KEYEVENTF_KEYUP, 0)  # 松开按键
    win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)
    time.sleep(1)
    win32api.mouse_event(win32con.MOUSEEVENTF_MOVE | win32con.MOUSEEVENTF_ABSOLUTE, 32768, 32768, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    win32api.keybd_event(13, 0, 0, 0)  # (回车)
    win32api.keybd_event(13, 0, win32con.KEYEVENTF_KEYUP, 0)  # 松开按键
    time.sleep(0.5)
    win32api.keybd_event(17, 0, 0, 0)  # (回车)
    win32api.keybd_event(65, 0, 0, 0)  # (回车)
    win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)
    win32api.keybd_event(65, 0, win32con.KEYEVENTF_KEYUP, 0)
    time.sleep(0.5)
    win32api.keybd_event(13, 0, 0, 0)  # (回车)
    win32api.keybd_event(13, 0, win32con.KEYEVENTF_KEYUP, 0)
    time.sleep(0.5)

def checkData(): #检查当前录入缴费的所需要的资料
    global period
    #e_xlsx = load_workbook(path+"非平峰谷电费缴费导入模板(软件用).xlsx")
    #s1 = e_xlsx['非平峰谷']
    s1 = template['非平峰谷']
    meter = s1['F2']
    percent = s1['H2']
    dateStr = s1['C2']
    print(meter.value)
    if percent.value==1 or percent.value=="1":
        print("独享")
        share="独享"
    else:
        print("共享")
        share = "共享"
    if type(dateStr.value) == datetime.datetime:
        period =int(dateStr.value.year)*100+dateStr.value.month+1
    if type(dateStr.value) == str:
        list = dateStr.value.split('-')  # 分割日期
        print(list[0])
        print(list[1])
        print(list[2])
        periodYear = str(list[0])
        periodMonth = str(int(list[1])+1).zfill(2)
        if periodYear == "2019" and list[1]=="12":
            period = 202001
        elif periodYear == "2018" and list[1]=="12":
            period = 201901
            periodYear="2019"
            periodMonth="01"
        else:
            period = str(list[0])+str(int(list[1])+1).zfill(2)
    print(period)
    #print(periodYear)
    #print(periodMonth)
    xlsx2 = load_workbook(path+"报账点数据.xlsx")

    sht=xlsx2['Sheet1']
    col = sht['C']
    #确定电表数据位置
    billNo = 'D'
    billDate = 'E'

    i = 0
    for cell in col:
        i += 1
        """
        if cell.value==meter.value and str(sht['E'+cell.coordinate[1:]].value)==str(period):
            regionLocat = 'A' + cell.coordinate[1:]
            billNo = 'D' + cell.coordinate[1:]
            billDate = 'E' + cell.coordinate[1:]
            break
        """
        if cell.value==meter.value :
            regionLocat = 'A' + cell.coordinate[1:]
            billNo = 'D' + cell.coordinate[1:]
            billDate = 'E' + cell.coordinate[1:]
            break

    region = sht[regionLocat] #获取报账点区域
    Date = sht[billDate]
    No = sht[billNo]
    print(region.value)
    print(No.value)
    print(Date.value)

    #
    global billPath

    if int(period) > 201901:
        filePath = "\""+path+"缴费所需资料\\"+str(period)+"直供电资料\\"+region.value+"\\"+share+"\""
    if int(period) <= 201901:
        filePath = "\"E:\Work\\0759\\QT-001-直供电" + share + "清单（电子版）.xlsx\" \"E:\Work\\0759\\QT-002-"+periodYear+"年"+periodMonth+"月供电局清单（电子版）.xlsx\""
        if share=="共享":
            filePath = "\"E:\Work\\0759\\QT-001-直供电" + share + "清单（电子版）.xlsx\" \"E:\Work\\0759\\QT-002-" + periodYear + "年" + periodMonth + "月供电局清单（电子版）.xlsx\" \"" + letterPath + "\""
            billPath = "\"E:\Work\\0759\\影像资料PDF\\"+periodYear+"-"+periodMonth+"影像资料\\直供电\\"+No.value+"\""
        if share=="独享":
            filePath = "\"E:\Work\\0759\\QT-001-直供电" + share + "清单（电子版）.xlsx\" \"E:\Work\\0759\\QT-002-" + periodYear + "年" + periodMonth + "月供电局清单（电子版）.xlsx\""
            billPath = "\"E:\Work\\0759\\影像资料PDF\\"+periodYear+"-"+periodMonth+"影像资料\\直供电\\"+No.value+"\""
    print(filePath)
    return filePath
    xlsx2.close()

def login(user): #登录
    element = browser.find_element_by_id("area_name")
    ActionChains(browser).move_to_element(element).perform()
    element.click()
    element = browser.find_element_by_xpath(
        "/html/body/div[3]/div/div[2]/form/div[1]/div[3]/div[3]/div/div[6]/div[3]/span[1]")
    ActionChains(browser).move_to_element(element).perform()
    element.click()  # 选择区域
    if user==userName1:
        element = browser.find_element_by_xpath("/html/body/div[3]/div/div[2]/form/div[1]/div[1]/input")
        element.send_keys(user1)
        element = browser.find_element_by_xpath("/html/body/div[3]/div/div[2]/form/div[2]/div/input[1]")
        element.send_keys(password1)
    if user==userName2:
        element = browser.find_element_by_xpath("/html/body/div[3]/div/div[2]/form/div[1]/div[1]/input")
        element.send_keys(user2)
        element = browser.find_element_by_xpath("/html/body/div[3]/div/div[2]/form/div[2]/div/input[1]")
        element.send_keys(password2)
    time.sleep(10)  # 有10秒的时间输入验证码
    element = browser.find_element_by_xpath("/html/body/div[3]/div/div[2]/form/div[5]/div/button")
    ActionChains(browser).move_to_element(element).perform()
    element.click() #点击登录按钮

def Import(): #批量导入功能的的主函数
    index = 9  # 检查上传成功按钮的xpath位置
    col = toolSheet['A']

    k = 0
    for cell in col:
        k += 1
        if cell.value == None:
            rowMax = cell.row - 1
            break #遍历A列获取有数据的行数

    i = 1
    for i in range(1, rowMax): #遍历列
        i += 1
        row = toolSheet[str(i)]
        j = 0
        for cell in row: #遍历行
            j += 1
            if j==column_index_from_string('AD') or j==column_index_from_string('AE')or j==column_index_from_string('AY')or j==column_index_from_string('BA')or j==column_index_from_string('BB')or j==column_index_from_string('BH')or j==column_index_from_string('BJ')or j==column_index_from_string('BK'):
                if cell.value == None:
                    data =" "
                else:
                    data = new_round(cell.value, 2) #可能因为出账工具的bug,某些时候获取金额到时候可能会出现一堆小数,此功能会自动四舍五入
            elif type(cell.value) == datetime.datetime:

                year = cell.value.strftime('%Y')
                month = cell.value.strftime('%m')
                day = cell.value.strftime('%d')
                #data=day + '/' + month + '/' + year

                datadate = datetime.date(int(year), int(month), int(day))
                #data = datadate.strftime("%#d/%#m/%Y")
                data = datadate.strftime("%Y-%#m-%#d") #去除日期前导0
                #data = cell.value.strftime("%#d/%#m/%Y") #涉及日期的处理 python的智能对象会识别为时间对象,这是转换成日期输出为yyyy-mm-dd的格式
            elif cell.value == None:
                data =' '
            else:
                data = cell.value  #空值None会被当成"None"字符串输出,这里把None转换成一个空字符

            """    
            elif cell.value == None and j >= column_index_from_string('AW'):
                data = ' '
            elif cell.value == None and j < column_index_from_string('AW'):
                data = ' '
            """

            templateSheet[str(get_column_letter(j)) + '2'].value = str(data) #写入导入模板的第二行
        template.save(path+"非平峰谷电费缴费导入模板(软件用).xlsx") #保存

        elecAccountImport(index)  # 导入电费

        index+=1 #每次导入电费后,每一次网页的提示元素都会往新增一个div标签来显示提示框,这里index值将传入下一次导入电费时候检索状态元素使用的Xpath路径

def elecAccountImport(index): #导入电费时的网页操作函数
    file = checkData()
    element = browser.find_element_by_xpath("/html/body/div[1]/button[4]")
    ActionChains(browser).move_to_element(element).perform()
    element.click()
    time.sleep(0.5)

    # 定位上传按钮，添加本地文件
    browser.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/form/div/div/input[1]").send_keys(path+"非平峰谷电费缴费导入模板(软件用).xlsx")
    time.sleep(0.5) #上传模板

    f1 = browser.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div/ul[1]/li/div[2]/a")
    UpLoad_File(f1,file) #非input标签上传文件 #上传附件
    time.sleep(0.5)
    element = browser.find_element_by_xpath("/html/body/div[7]/div/div/div[3]/a")
    ActionChains(browser).move_to_element(element).perform()
    element.click()
    time.sleep(0.5)
    if int(period) <= 201901:
        f1 = browser.find_element_by_xpath("/html/body/div[4]/div/div/div[2]/div/ul[1]/li/div[2]/a")
        UpLoad_File(f1, billPath)  # 非input标签上传文件 #上传附件
        time.sleep(1)
        element = browser.find_element_by_xpath("/html/body/div[7]/div/div/div[3]/a")
        ActionChains(browser).move_to_element(element).perform()
        element.click()
        time.sleep(1)
    element = browser.find_element_by_xpath("/html/body/div[4]/div/div/div[3]/a[2]")
    ActionChains(browser).move_to_element(element).perform()
    element.click()
    time.sleep(0.2)

    while 1:
        try:
            text = browser.find_element_by_xpath("/html/body/div[" + str(index) + "]/div/div/div[2]").text
            print(text)
            break
        except:
            continue #检测导入完成的功能呢

    element = browser.find_element_by_xpath("/html/body/div[" + str(index) + "]/div/div/div[3]/button")
    ActionChains(browser).move_to_element(element).perform()
    element.click() #导入完成后点击完成按钮
    print(index-9+1)

def readConfig():
    global path,letterPath,userName1,user1,password1,userName2,user2,password2
    config = get_yaml_data('config.yaml')
    path = config['path']
    letterPath = config['letterPath']

    userName1 = config['userName1']
    user1 = config['user1']
    password1 = config['password1']

    userName2 = config['userName2']
    user2 = config['user2']
    password2 = config['password2']

def benchmark():
    xlsx1 = load_workbook(path + "标杆数据.xlsx")
    s1 = xlsx1['Sheet1']
    col1 = s1['A']
    q = queue.Queue()
    url1 = "http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment/showBenchmark"  # 接口地址
    # js = "window.open('http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment_finance/manage.html?formData=%7B%22billamountDateOpen%22:%222020-08-13%22%7D')"
    js = "window.open('http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment_finance/manage.html')"
    browser.execute_script(js)
    time.sleep(0.5)
    page = browser.window_handles
    browser.switch_to_window(page[1])  # 切换至明细页面
    time.sleep(1)
    flag = input("请调整好爬取的页面,按1继续\n>>>");
    if flag == "1":
        i = 1
        for link in browser.find_elements_by_xpath("//tbody//*[@href]"):  # 此循环请求速度极快 有被服务器ban掉的可能
            accountLink = link.get_attribute('href')
            accountId = re.findall(r"billaccountpaymentdetailId=(.+?)&", accountLink, flags=0)
            # 消息头数据
            headers = {
                'Connection': 'keep-alive',
                'Content-Length': '59',
                'Accept': '*/*',
                'DNT': '1',
                'X-Requested-With': 'XMLHttpRequest',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36',
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                'Origin': 'http://10.217.240.219:8090',
                'Referer': accountLink,
                'Accept-Encoding': 'gzip, deflate',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Cookie': '/NCMS/welcomecuNum=1; SESSION=' + cookie['value'],

            }
            # 消息数据
            data = {'billaccountpaymentdetailId': accountId[0]}
            r = requests.post(url1, headers=headers, data=data, verify=False)  # 发送POST请求标杆
            packet = r.json()  # 获取回复
            html = packet['obj']  # 提取html
            """
            try:
                soup = BeautifulSoup(html, 'html.parser')  # 将html实例化为BeautifulSoup对象
            except:
                print("html实例化失败,检查你的cookie是否正常获取")
            """
            soup = BeautifulSoup(html, 'html.parser')  # 将html实例化为BeautifulSoup对象
            for item in soup.find_all("td"):  # 遍历html提取table
                benchmark = re.sub('\s', ' ', item.get_text())  # 去除html标签
                print(benchmark)
                q.put(benchmark)  # 数值推入队列
            account = browser.find_element_by_xpath(
                "/html/body/div[4]/div[2]/div[2]/table/tbody/tr[" + str(i) + "]/td[7]").text  # 获取缴费单号
            accountName = browser.find_element_by_xpath(
                "/html/body/div[4]/div[2]/div[2]/table/tbody/tr[" + str(i) + "]/td[8]").text  # 获取缴费单号
            accountDate = browser.find_element_by_xpath(
                "/html/body/div[4]/div[2]/div[2]/table/tbody/tr[" + str(i) + "]/td[11]").text  # 获取缴费单号
            DateStr = browser.find_element_by_xpath(
                "/html/body/div[4]/div[2]/div[2]/table/tbody/tr[" + str(i) + "]/td[12]").text  # 获取缴费单号
            DateEnd = browser.find_element_by_xpath(
                "/html/body/div[4]/div[2]/div[2]/table/tbody/tr[" + str(i) + "]/td[13]").text  # 获取缴费单号
            print(account)
            k = 1
            for k in range((i * 4 - 3), (i * 4 + 1)):  # 遍历列
                k += 1
                s1['E' + str(k)].value = str(account)
                s1['F' + str(k)].value = str(accountName)
                s1['G' + str(k)].value = str(accountDate)
                s1['H' + str(k)].value = str(DateStr)
                s1['I' + str(k)].value = str(DateEnd)
                row = s1[str(k)]
                j = 0
                for cell in row:  # 遍历行
                    j += 1
                    if j == 5:
                        break
                    cell.value = q.get()
            i += 1
        xlsx1.save(path + "标杆数据.xlsx")

def histoicFlowTime():
    url1 = "http://10.217.240.219:8090/NCMS/act/common/histoicFlowTime"  # 接口地址
    js = "window.open('http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment_finance/manage.html')"
    browser.execute_script(js)
    time.sleep(0.5)
    page = browser.window_handles
    browser.switch_to_window(page[1])  # 切换至明细页面
    time.sleep(1)
    user = input("请调整好爬取的页面,按1继续\n>>>");
    if user == "1":
        i = 1
        for link in browser.find_elements_by_xpath("//tbody//*[@href]"):  # 此循环请求速度极快 有被服务器ban掉的可能
            accountLink = link.get_attribute('href')
            accountId = re.findall(r"billaccountpaymentdetailId=(.+?)&", accountLink, flags=0)
            # 消息头数据
            headers = {
                'Connection': 'keep-alive',
                'Content-Length': '57',
                'Accept': 'application/json, text/javascript, */*; q=0.01',
                'DNT': '1',
                'X-Requested-With': 'XMLHttpRequest',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36',
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                'Origin': 'http://10.217.240.219:8090',
                'Referer': accountLink,
                'Accept-Encoding': 'gzip, deflate',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Cookie': '/NCMS/welcomecuNum=1; SESSION=' + cookie['value'],

            }
            # 消息数据
            data = {
                'tableName': 'ele_payment',
                'id': accountId[0],
            }

            r = requests.post(url1, headers=headers, data=data, verify=False)  # 发送POST请求标杆
            packet = r.json()  # 获取回复
            obj = str(packet['obj'])
            str_info=obj.strip( '[]' )
            #str_info = test1.replace("\'", "\"")
            dict_info = eval(str_info)
            dataMax=len(dict_info)
            print(eval(str(dict_info[dataMax-2]))['comment'])

def BillaccountPaymentInfo():
    url1 = "http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment/queryVEleBillaccountPaymentInfo"  # 接口地址
    # 消息头数据
    headers = {
        'Connection': 'keep-alive',
        'Content-Length': '297',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'DNT': '1',
        'X-Requested-With': 'XMLHttpRequest',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Origin': 'http://10.217.240.219:8090',
        'Referer': 'http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment_finance/manage.html',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cookie': '/NCMS/welcomecuNum=1; SESSION=' + cookie['value'],
    }
    # 消息数据
    data = {
        'pageNumber': '1',  # 页码 (暂时没找到获取有多少页的接口)
        'pageSize': '10',  # 每页显示多少条记录 (10 25 50 100 500)
        'billaccountCode': '',  # 报账点编码
        'billaccountName': '',  # 报账点名称
        'paymentCode': '',  # 缴费单编码
        'pregId': '',  # 地市 湛江=36b7ee0a05d04884b332effa938fc682 此字符串采用MD5加密需要暴力穷举解密
        'regId': '',
        # 区县 徐闻=15dfd3edcd61487793ccf54e0cabf13b 霞山=1cb22e96e86e462d91fa369581fd602b 遂溪=2e36821772684b95a93a34a1921c3816 开发区=49a83c48cd794270baa78f2b311fc804 坡头=70b4bb88dc6c4fc9bc32b5cceee43207 吴川=99a7b6b3efd348c8afca9f4e40d4219a 雷州=e6cf859af33547338281d710a623dabb 赤坎=f8707152932f4186b322bcf89523ab94 麻章=fa0c8d2af1474accaf022465376152a6 廉江=fbdc86b3ae0d4916afb5f031f1e74de8
        'auditingState': '',  # 审批状态 审核通过=0 审核未通过=8 审核中=9 未提交=-1
        'supplyMethod': '',  # 供电方式 直供电=1 转供电=2
        'billaccountType': '2',  # 报账点类型自维报账点=1 铁塔电费报账点=2 代持报账点=3
        'startdateOpen': '',  # 缴费期始 YYYY-MM-DD
        'startdateClose': '',  # 缴费期始 YYYY-MM-DD
        'enddateOpen': '',  # 缴费期终 YYYY-MM-DD
        'enddateClose': '',  # 缴费期终 YYYY-MM-DD
        'userCodeOrName': '',  # 录入人 梁湛波=%E6%A2%81%E6%B9%9B%E6%B3%A2  此字符串采用URL UTF-8编码 可自行解码
        'billamountDateOpen': '',  # 缴费申请日期 YYYY-MM-DD
        'billamountDateClose': '',  # 缴费申请日期 YYYY-MM-DD
        'overflow': '',  # 是否超标 不超标=0 超标=1
        'isFinance': '1',  # 不太清楚这个选项有什么用
        'billamountState': '',  # 推送状态
        'dateSort': '1',  # 排列方式 缴费申请日期从新到旧=1 缴费申请日期从旧到新=2 缴费期始从新到旧=3 缴费期始从旧到新=4 缴费期终从新到旧=5 缴费期终从旧到新=6
    }

    r = requests.post(url1, headers=headers, data=data, verify=False)  # 发送POST请求标杆
    packet = r.json()  # 获取回复
    obj = str(packet['obj'])
    str_info = obj.strip('[]')
    dict_info = eval(str_info)
    str_info=str(dict_info['list'])
    billList = eval(str_info.strip('[]'))
    billInfo=eval(str(billList[0]))
    print(billInfo)
    print(billInfo['paymentdetailNote'])

def ElectricmeterInfoDetail():
    url1 = "http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment/queryElectricmeterInfoDetail"  # 接口地址
    # 消息头数据
    headers = {
        'Connection': 'keep-alive',
        'Content-Length': '297',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'DNT': '1',
        'X-Requested-With': 'XMLHttpRequest',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Origin': 'http://10.217.240.219:8090',
        'Referer': 'http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment_finance/manage.html',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cookie': '/NCMS/welcomecuNum=1; SESSION=' + cookie['value'],
    }
    # 消息数据
    data = {
        'billaccountpaymentdetailId':'869e1e6ac4a540678c108c07da01c867',
        'isFinance': '1'
    }

    r = requests.post(url1, headers=headers, data=data, verify=False)  # 发送POST请求标杆
    packet = r.json()  # 获取回复
    obj = str(packet['obj'])
    str_info = obj.strip('[]')
    dict_info = eval(str_info)
    billInfo=eval(str(dict_info['contract']))
    otherCost = eval(str(billInfo['otherCost']).strip('[]'))
    meterInfo=eval(str(dict_info['meterlist']).strip('[]'))
    #resourceInfo = eval(str(dict_info['resourceList']).strip('[]'))
    print(billInfo)#缴费明细
    print(otherCost)#其他费用
    print(meterInfo)#电表缴费明细
    #print(resourceInfo)

if __name__ == '__main__':
    try:
        readConfig()
    except:
        print("读取配置文件失败")
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    browser = webdriver.Chrome(options=options)
    # browser = webdriver.Chrome()
    url = 'http://10.217.240.219:8090/NCMS/welcome'
    # 初次建立连接，随后方可修改cookie
    browser.get(url)
    # 删除第一次建立连接时的cooki
    browser.delete_all_cookies()
    # 读取登录时存储到本地的cookie
    try:
        with open('cookies.json', 'r', encoding='utf-8') as f:
            listCookies = json.loads(f.read())
        for cookie in listCookies:
            browser.add_cookie({
                'domain': '10.217.240.219',  # 此处xxx.com前，需要带点
                'name': cookie['name'],
                'value': cookie['value'],
                'path': '/',
                'expires': None
            })
        browser.get(url)
    except:
        browser.get(url)
    # 再次访问页面，便可实现免登陆访问
    while 1:
        print("(1) - 直供电电费导入(在出账工具出账后使用)")
        print("(2) - 登录,获取cookie")
        print("(3) - 报账点批量启用-(不可靠)")
        print("(4) - 超标备注(不可靠)")
        print("(5) - 爬取标杆(beta)")
        print("(6) - 爬取审核信息(beta)")
        print("(7) - 爬取电费列表(beta)")
        print("(8) - 爬取电费明细(beta)")
        print("(0) - 退出")
        userInput = input("输入数字选择功能\n>>>");
        if userInput == "0":
            break
        if userInput == "1":
            js = "window.open('http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/payment_finance/record.html')"
            browser.execute_script(js)
            page = browser.window_handles
            browser.switch_to_window(page[1])  # 切换至缴费页面
            try:
                tool = load_workbook(path+"三家运营商代垫电费出账工具V35V1-20200818.xlsm")
                template = load_workbook(path+"非平峰谷电费缴费导入模板(软件用).xlsx")
                toolSheet = tool['新版导入表']
                templateSheet = template['非平峰谷']
            except:
                print("文档被占用,请解除占用后使用")
            Import()
            tool.close()
            template.close()
            print("导入完成!")
            browser.close()
        if userInput == "2":
            chooseUser = input(userName1+"-(1) "+userName2+"-(2) 返回-(0)\n")
            print("有10秒的时间输入验证码,输入后不要操作任何按钮,10秒后自动登录")
            if chooseUser=="1":
                login(userName1)
            if chooseUser == "2":
                login(userName2)
            if chooseUser == "0":
                continue
            dictCookies = browser.get_cookies()
            jsonCookies = json.dumps(dictCookies)
            # 登录完成后，将cookie保存到本地文件
            with open('cookies.json', 'w') as f:
                f.write(jsonCookies)
            print("已获取cookie 请重启程序...")
            sys.exit()
        if userInput == "3":
            js = "window.open('http://10.217.240.219:8090/NCMS/asserts/tpl/selfelec/billaccount/manage.html')"
            browser.execute_script(js)
            time.sleep(0.5)
            page = browser.window_handles
            browser.switch_to_window(page[1])  # 切换至缴费页面

            xlsx= load_workbook("报账点.xlsx")

            Sheet = xlsx['Sheet1']

            col = Sheet['A']

            time.sleep(5)

            i = 0
            for cell in col:
                i += 1
                point=cell.value
                print(point)

                browser.find_element_by_xpath("/html/body/form/div[1]/input").clear()

                browser.find_element_by_xpath("/html/body/form/div[1]/input").send_keys(point)

                element = browser.find_element_by_xpath("/html/body/form/button")
                ActionChains(browser).move_to_element(element).perform()
                element.click()
                time.sleep(8)

                element = browser.find_element_by_xpath("/html/body/div[3]/div[2]/div[2]/table/tbody/tr/td[1]/input")
                ActionChains(browser).move_to_element(element).perform()
                element.click()
                time.sleep(1)

                element = browser.find_element_by_xpath("/html/body/div[2]/button[6]")
                ActionChains(browser).move_to_element(element).perform()
                element.click()
                time.sleep(1)

                element = browser.find_element_by_xpath("/html/body/div[7]/div/div/div[2]/div[1]/input")
                ActionChains(browser).move_to_element(element).perform()
                element.click()
                time.sleep(1)

                element = browser.find_element_by_xpath("/html/body/div[7]/div/div/div[3]/button[2]")
                ActionChains(browser).move_to_element(element).perform()
                element.click()
                time.sleep(6)
                win32api.mouse_event(win32con.MOUSEEVENTF_MOVE | win32con.MOUSEEVENTF_ABSOLUTE, 32768, 32768, 0, 0)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        if userInput == "4":
            billNo = input("输入供电用户号>>>");
            xlsx1 = load_workbook(path + "东海区域2020年6月直供电电子版清单.xlsx")
            xlsx2 = load_workbook(path + "超标数据.xlsx")
            s1 = xlsx1['202006']
            s2 = xlsx2['Sheet1']
            col1 = s1['B']
            col2 = s2['D']
            i = 0
            test=0
            for cell in col1:
                i += 1
                if str(cell.value) == str(billNo):
                    test = s1['A' + cell.coordinate[1:]].value
                    print(s1['A' + cell.coordinate[1:]].value)
            j = 0
            for cell in col2:
                j += 1
                if str(cell.value) == test:
                    print(s2['C' + cell.coordinate[1:]].value)
        if userInput == "5":
            benchmark()
        if userInput == '6':
            histoicFlowTime()
        if userInput == '7':
            BillaccountPaymentInfo()
        if userInput == '8':
            ElectricmeterInfoDetail()
        if userInput == "V" or userInput == "v":
            print(version)
    browser.switch_to_window(page[0])
    browser.close()
    sys.exit(0)
